import xml.etree.ElementTree as ET
import openpyxl as xls


def proc_mallas(data_path, output_path):
    # activate excel workbook and properties
    wb = xls.Workbook()
    sheet = wb.active
    rowJobName = 1
    rowData = 2
    column_key = 1
    column_value = 2

    # reading xml file
    print('Abriendo archivo xml')
    tree = ET.parse(data_path)
    root = tree.getroot()

    # processing xml file
    print('Procesando información xml --> -->')
    for folder in root:
        for key, value in folder.attrib.items():
            if key == 'FOLDER_NAME':
                sheet.cell(rowJobName, column_key, key)
                sheet.cell(rowJobName, column_value, value)
        for jobs in folder:
            for key, value in jobs.attrib.items():
                if key == 'APPLICATION':
                    sheet.cell(rowData, column_key, "-------")
                    rowData += 1
                    sheet.cell(rowData, column_key, key)
                    sheet.cell(rowData, column_value, value)
                    rowData += 1
                if key == 'CMDLINE':
                    sheet.cell(rowData, column_key, key)
                    sheet.cell(rowData, column_value, value)
                    rowData += 1
                if key == 'JOBNAME':
                    sheet.cell(rowData, column_key, key)
                    sheet.cell(rowData, column_value, value)
                    rowData += 1
                if key == 'DESCRIPTION':
                    sheet.cell(rowData, column_key, key)
                    sheet.cell(rowData, column_value, value)
                    rowData += 1
                if key == 'RUN_AS':
                    sheet.cell(rowData, column_key, key)
                    sheet.cell(rowData, column_value, value)
                    rowData += 1
                if key == 'NODEID':
                    sheet.cell(rowData, column_key, key)
                    sheet.cell(rowData, column_value, value)
                    rowData += 1
                if key == 'CRITICAL':
                    sheet.cell(rowData, column_key, key)
                    sheet.cell(rowData, column_value, value)
                    rowData += 1
                if key == 'CREATION_DATE':
                    sheet.cell(rowData, column_key, key)
                    sheet.cell(rowData, column_value, value)
                    rowData += 1
                if key == 'MEMLIB':
                    sheet.cell(rowData, column_key, key)
                    sheet.cell(rowData, column_value, value)
                    rowData += 1
                if key == 'MEMNAME':
                    sheet.cell(rowData, column_key, key)
                    sheet.cell(rowData, column_value, value)
                    rowData += 1
                if key == 'SUB_APPLICATION':
                    sheet.cell(rowData, column_key, key)
                    sheet.cell(rowData, column_value, value)
                    rowData += 1
                if key == 'TIMEFROM':
                    sheet.cell(rowData, column_key, key)
                    sheet.cell(rowData, column_value, value)
                    rowData += 1

    wb.save(output_path)
    print('** Archivo excel creado con exito **')


mesh_path = '/home/edgar-gonzalez/Escritorio/desarrollos/python/DataProcessingMesh/data/CR-MXSFRDIA-T02.xml'
output_path = 'data.xlsx'
proc_mallas(mesh_path, output_path)
